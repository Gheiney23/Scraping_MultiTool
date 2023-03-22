import pandas as pd
import time
import re
import os
import PySimpleGUI as sg
from urllib.request import urlretrieve
from openpyxl import load_workbook
from selenium import webdriver as wb
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl.worksheet.properties import WorksheetProperties as wp
from selenium.webdriver.common.action_chains import ActionChains
from urllib.error import HTTPError

def img_tool(sku_list, uid_list, f_id_list, file_name):
        
    # Setting up the webdriver for Selenium
    options = wb.ChromeOptions()
    options.add_argument('--start-maximized')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    driver = wb.Chrome(options=options)

    src_dict = {'Sku': [], 'Img_url1': [], 'Img_url2': [], 'Img_url3': [], 'Img_url4': [], 'Img_url5': [], 'Img_url6': [], 'Img_url7': [], 'Img_url8': [], 'Img_url9': [], 'Skus_Not_Found': []}
    
    for (id1, id2, sku) in zip(uid_list, f_id_list, sku_list):
        try:
            path = 'url_here_id1{}?id2={}'.format(id1, id2)
            driver.get(path)
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "root")))
            # time.sleep(1)
            driver.execute_script("window.scrollTo(0, 300)")
            
            # Extracting the image src
            src_1 = driver.find_element_by_xpath("//*[contains(@class,'w-auto self-center undefined')]").get_attribute('src')
            # time.sleep(1)
            src_dict['Img_url1'].append(src_1)
            src_dict['Sku'].append(sku)
            
            try:
                element = driver.find_element_by_xpath("//div[@aria-label='thumb slide 1']")
                actions = ActionChains(driver)
                actions.move_to_element(element).perform()
                element.click()
                time.sleep(1)
                src_2 = driver.find_element_by_xpath("//div[starts-with(@aria-label, '2 /')]").find_element_by_tag_name('img').get_attribute('src')
                src_dict['Img_url2'].append(src_2)
            except:
                src_dict['Img_url2'].append('NULL')
                
            try:    
                element = driver.find_element_by_xpath("//div[@aria-label='thumb slide 2']")
                actions = ActionChains(driver)
                actions.move_to_element(element).perform()
                element.click()
                time.sleep(1)
                src_3 = driver.find_element_by_xpath("//div[starts-with(@aria-label, '3 /')]").find_element_by_tag_name('img').get_attribute('src')
                src_dict['Img_url3'].append(src_3)
            except:
                src_dict['Img_url3'].append('NULL')

            try:   
                element = driver.find_element_by_xpath("//div[@aria-label='thumb slide 3']")
                actions = ActionChains(driver)
                actions.move_to_element(element).perform()
                element.click()
                time.sleep(1)
                src_4 = driver.find_element_by_xpath("//div[starts-with(@aria-label, '4 /')]").find_element_by_tag_name('img').get_attribute('src')
                src_dict['Img_url4'].append(src_4)
            except:
                src_dict['Img_url4'].append('NULL')

            try:   
                element = driver.find_element_by_xpath("//div[@aria-label='thumb slide 4']")
                actions = ActionChains(driver)
                actions.move_to_element(element).perform()
                element.click()
                time.sleep(1)
                src_5 = driver.find_element_by_xpath("//div[starts-with(@aria-label, '5 /')]").find_element_by_tag_name('img').get_attribute('src')
                src_dict['Img_url5'].append(src_5)
            except:
                src_dict['Img_url5'].append('NULL')

            try:   
                element = driver.find_element_by_xpath("//div[@aria-label='thumb slide 5']")
                actions = ActionChains(driver)
                actions.move_to_element(element).perform()
                element.click()
                time.sleep(1)
                src_6 = driver.find_element_by_xpath("//div[starts-with(@aria-label, '6 /')]").find_element_by_tag_name('img').get_attribute('src')
                src_dict['Img_url6'].append(src_6)
            except:
                src_dict['Img_url6'].append('NULL')
            
            try:   
                element = driver.find_element_by_xpath("//div[@aria-label='thumb slide 6']")
                actions = ActionChains(driver)
                actions.move_to_element(element).perform()
                element.click()
                time.sleep(1)
                src_7 = driver.find_element_by_xpath("//div[starts-with(@aria-label, '7 /')]").find_element_by_tag_name('img').get_attribute('src')
                src_dict['Img_url7'].append(src_7)
            except:
                src_dict['Img_url7'].append('NULL')
            
            try:   
                element = driver.find_element_by_xpath("//div[@aria-label='thumb slide 7']")
                actions = ActionChains(driver)
                actions.move_to_element(element).perform()
                element.click()
                time.sleep(1)
                src_8 = driver.find_element_by_xpath("//div[starts-with(@aria-label, '8 /')]").find_element_by_tag_name('img').get_attribute('src')
                src_dict['Img_url8'].append(src_8)
            except:
                src_dict['Img_url8'].append('NULL')
            
            try:   
                element = driver.find_element_by_xpath("//div[@aria-label='thumb slide 8']")
                actions = ActionChains(driver)
                actions.move_to_element(element).perform()
                element.click()
                time.sleep(1)
                src_9 = driver.find_element_by_xpath("//div[starts-with(@aria-label, '9 /')]").find_element_by_tag_name('img').get_attribute('src')
                src_dict['Img_url9'].append(src_9)
            except:
                src_dict['Img_url9'].append('NULL')
        
        except:
            src_dict['Sku'].append(sku)       
            src_dict['Skus_Not_Found'].append(sku)
            src_dict['Img_url1'].append('NULL')
            src_dict['Img_url2'].append('NULL')
            src_dict['Img_url3'].append('NULL')
            src_dict['Img_url4'].append('NULL')
            src_dict['Img_url5'].append('NULL')
            src_dict['Img_url6'].append('NULL')
            src_dict['Img_url7'].append('NULL')
            src_dict['Img_url8'].append('NULL')
            src_dict['Img_url9'].append('NULL')

    # quitting the driver and manipulation the dictionary into a dataframe
    driver.quit()

    df = pd.DataFrame.from_dict(src_dict,orient='index')
    df = df.transpose()
    # # df['Img_url'].fillna('NULL', inplace=True)

    # Writing the dataframe to an excel worksheet
    path = r'{}'.format(file_name)
    excel_wb = load_workbook(path)
    with pd.ExcelWriter(path) as writer:
        writer.book = excel_wb
        df.to_excel(writer, sheet_name='Asset_Data', index=False)
        file_sheet = writer.sheets['Asset_Data']
        file_sheet.sheet_properties.tabColor = 'FFFF00'

def bullet_tool(sku_list, uid_list, f_id_list, file_name):
    
    data_dict = {'Sku': [], 'Bullet1': [], 'Bullet2': [], 'Bullet3': [], 'Bullet4': [], 'Bullet5': [], 'Bullet6' : [], 'Bullet7': [], 'Bullet8': [], 'Bullet9': [], 'Skus_Not_Found': []}

    # Setting up the webdriver for Selenium
    options = wb.ChromeOptions()
    options.add_argument('--start-maximized')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    driver = wb.Chrome(options=options)

    # Iterating through all lists and injecting relevant data into the url
    for (id1, id2, sku) in zip(uid_list, f_id_list, sku_list):
        path = 'url_here_id1{}?id2={}'.format(id1, id2)
        driver.get(path)
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "root")))
        # time.sleep(1)

        # If the sku is located
        try:
            
            # Removal of banner if necessary
            # if driver.find_element_by_xpath("//div[contains(@class, 'ku8y0w3')]"):
            #     driver.find_element_by_xpath("//div[contains(@class, 'ku8y0w6')]").click()
            # else:
            #     continue
            
            # Moving to the bullet points on the webpage
            element = driver.find_element_by_xpath("//*[@id='main-content']/div/section[4]/div[1]/section/div[1]/section/div[3]/section/div/div[1]/div[1]/ul[1]")
            actions = ActionChains(driver)
            actions.move_to_element(element).perform()
            time.sleep(1)
            
            # Extracting the web element then transforming it to text with no special characters
            li_elements = element.find_elements_by_tag_name('li')
            b_list = []
            
            for li in li_elements:
                li_text = li.text
                bullet = re.sub("[^A-Za-z0-9 -\/]", "", li_text)
                bullet = bullet.replace('"', "-in") 
                b_list.append(bullet)
            
            # Loading all text found into the data_dict
            try:
                data_dict['Bullet1'].append(b_list[0])
            except:
                data_dict['Bullet1'].append('NULL')
            
            try:
                data_dict['Bullet2'].append(b_list[1])
            except:
                data_dict['Bullet2'].append('NULL')
            
            try:
                data_dict['Bullet3'].append(b_list[2])
            except:
                data_dict['Bullet3'].append('NULL')
            
            try:
                data_dict['Bullet4'].append(b_list[3])
            except:
                data_dict['Bullet4'].append('NULL')

            try:
                data_dict['Bullet5'].append(b_list[4])
            except:
                data_dict['Bullet5'].append('NULL')
            
            try:
                data_dict['Bullet6'].append(b_list[5])
            except:
                data_dict['Bullet6'].append('NULL')
            
            try:
                data_dict['Bullet7'].append(b_list[6])
            except:
                data_dict['Bullet7'].append('NULL')
            
            try:
                data_dict['Bullet8'].append(b_list[7])
            except:
                data_dict['Bullet8'].append('NULL')
            
            try:
                data_dict['Bullet9'].append(b_list[8])
            except:
                data_dict['Bullet9'].append('NULL')

            data_dict['Sku'].append(sku)
                
        # If sku is not found
        except:
            data_dict['Sku'].append(sku)
            data_dict['Bullet1'].append('NULL')
            data_dict['Bullet2'].append('NULL')
            data_dict['Bullet3'].append('NULL')
            data_dict['Bullet4'].append('NULL')
            data_dict['Bullet5'].append('NULL')
            data_dict['Bullet6'].append('NULL')
            data_dict['Bullet7'].append('NULL')
            data_dict['Bullet8'].append('NULL')
            data_dict['Bullet9'].append('NULL')
            data_dict['Skus_Not_Found'].append(sku) 

    # Quitting the driver and creating a dataframe from data_dict   
    driver.quit()
    df = pd.DataFrame.from_dict(data_dict, orient='index')
    df = df.transpose()

    # Writing the dataframe to an excel worksheet
    path = r'{}'.format(file_name)
    excel_wb = load_workbook(path)
    with pd.ExcelWriter(path) as writer:
        writer.book = excel_wb
        df.to_excel(writer, sheet_name='Bullet_Data', index=False)
        file_sheet = writer.sheets['Bullet_Data']
        file_sheet.sheet_properties.tabColor = 'FFFF00'

def converter_tool(mfg_list_primary, mfg_list_2, mfg_list_3, mfg_list_4, Primary_list, img_2_list, img_3_list, img_4_list, excel_file_name, folder_name):
    # Create a dictionary from the two lists for a loop
    primary_img_dict = {mfg_list_primary[i]: Primary_list[i] for i in range(len(Primary_list))}
    img_2_dict = {mfg_list_2[i]: img_2_list[i] for i in range(len(img_2_list))}
    img_3_dict = {mfg_list_3[i]: img_3_list[i] for i in range(len(img_3_list))}
    img_4_dict = {mfg_list_4[i]: img_4_list[i] for i in range(len(img_4_list))}


    # Creating a folder variable for output
    output_directory = '{}'.format(folder_name)

    # Looping through the dictionary and creating .jpgs from the urls and loading the file names into a list
    file_name_dict = {'Primary_File_name':[], 'Image_2_Name': [], 'Image_3_Name': [], 'Image_4_Name': [], '404_error_images': []}


    for mfg, url in primary_img_dict.items():
        try:
            primary_file_name = mfg + '_Primary.jpg'
            urlretrieve(url, output_directory + f"\{primary_file_name}")
            file_name_dict['Primary_File_name'].append(primary_file_name)

        except HTTPError as err:
            if err.code == 404:
                file_name_dict['404_error_images'].append(mfg)
                pass
            else:
                raise
        
        # except:
        #     file_name_dict['Primary_File_name'].append('NULL')
        #     file_name_dict['Sku'].append(mfg)

    for mfg, url in img_2_dict.items():
        try:
            file_name2 = mfg + '_img2.jpg'
            urlretrieve(url, output_directory + f"\{file_name2}")
            file_name_dict['Image_2_Name'].append(file_name2)

        except HTTPError as err:
            if err.code == 404:
                file_name_dict['404_error_images'].append(mfg)
                pass
            else:
                raise
        
        # except:
        #     file_name_dict['Image_2_Name'].append('NULL')
        #     file_name_dict['Sku'].append(mfg)

    for mfg, url in img_3_dict.items():
        try:
            file_name3 = mfg + '_img3.jpg'
            urlretrieve(url, output_directory + f"\{file_name3}")
            file_name_dict['Image_3_Name'].append(file_name3)

        except HTTPError as err:
            if err.code == 404:
                file_name_dict['404_error_images'].append(mfg)
                pass
            else:
                raise
        
        # except:
        #     file_name_dict['Image_3_Name'].append('NULL')
        #     file_name_dict['Sku'].append(mfg)

    for mfg, url in img_4_dict.items():
        try:
            file_name4 = mfg + '_img4.jpg'
            urlretrieve(url, output_directory + f"\{file_name4}")
            file_name_dict['Image_4_Name'].append(file_name4)

        except HTTPError as err:
            if err.code == 404:
                file_name_dict['404_error_images'].append(mfg)
                pass
            else:
                raise
        
        # except:
        #     file_name_dict['Image_4_Name'].append('NULL')
        #     file_name_dict['Sku'].append(mfg)

    # Creating a dataframe from the file name list
    file_df = pd.DataFrame.from_dict(file_name_dict, orient='index')
    file_df = file_df.transpose()
    

    #  Writing the dataframe to an excel worksheet
    path = '{}'.format(excel_file_name)
    excel_wb = load_workbook(path)
    with pd.ExcelWriter(path) as writer:
        writer.book = excel_wb
        file_df.to_excel(writer, sheet_name='File_Data', index=False)
        file_sheet = writer.sheets['File_Data']
        file_sheet.sheet_properties.tabColor = 'FFFF00'

    sg.popup("Run Complete!")

def make_main_window():
    
    # Theme of windows
    sg.theme('Dark Grey 13')
    
    # Creating window layouts
    main_layout = [[sg.Text("Asset Scraper Tool")], 
                    [sg.Text("Choose which tool you want.")],
                    [sg.Button("Image URL Converter"), sg.Button("Image Scraper Tool"), sg.Button("Bullet Scraper Tool"), sg.Button("Exit")]]

    return sg.Window('Image Scraper Window', main_layout)

def make_img_window():
     # Theme of windows
    sg.theme('Dark Grey 13')

    img_layout = [[sg.Text("Image Scraper Tool")],
                  [sg.Text("Be sure to close the Excel file being used BEFORE running the tool.", text_color='red', font=('Arial Bold', 10))],
                  [sg.Text('Please enter Sku(MFG Number) list.'), sg.InputText(key='-SKU-', pad=(0,0))],
                  [sg.Text('Please enter ID1 list.'), sg.InputText(key='-UID-', pad=(0,0))],
                  [sg.Text('Please enter ID2 list.'), sg.InputText(key='-FID-', pad=(0,0))],
                  [sg.Text('Please enter the absolute path of Excel file to use.'), sg.InputText(key='-E_NAME-')],
                  [sg.Button("Run"), sg.Button("Exit")]]

    image_window = sg.Window('Image Scraper Window', img_layout, modal=True)

    while True:
        
        event, values = image_window.read()
        
        if event in(sg.WIN_CLOSED, "Exit"):
            break
        
        sku_list = values['-SKU-'].split('\n')
        uid_list = values['-UID-'].split('\n')
        f_id_list = values['-FID-'].split('\n')
        file_name = values['-E_NAME-'].rstrip()
        
        if event == 'Run':
            
            try:
                img_tool(sku_list, uid_list, f_id_list, file_name)
                sg.popup("Run Complete!")
            except:
                sg.popup("Something went wrong. Please make sure everything was entered correctly.")

    image_window.close()

def make_bullet_window():
    
    # Theme of windows
    sg.theme('Dark Grey 13')

    bullet_layout = [[sg.Text("Bullet Scraper Tool")],
                        [sg.Text("Be sure to close the Excel file being used BEFORE running the tool.", text_color='red', font=('Arial Bold', 10))],
                        [sg.Text('Please enter Sku(MFG Number) list.'), sg.InputText(key='-SKU-', pad=(0,0))],
                        [sg.Text('Please enter ID1 list.'), sg.InputText(key='-UID-', pad=(0,0))],
                        [sg.Text('Please enter ID2 list.'), sg.InputText(key='-FID-', pad=(0,0))],
                        [sg.Text('Please enter the absolute path of Excel file to use.'), sg.InputText(key='-E_NAME-')],
                        [sg.Button("Run"), sg.Button("Exit")]]
    
    bullet_window = sg.Window('Bullet Scraper Window', bullet_layout, modal=True)

    while True:
        
        event, values = bullet_window.read()
        
        if event in(sg.WIN_CLOSED, "Exit"):
            break
        
        sku_list = values['-SKU-'].split('\n')
        uid_list = values['-UID-'].split('\n')
        f_id_list = values['-FID-'].split('\n')
        file_name = values['-E_NAME-'].rstrip()

        if event == 'Run':
            
            try:
                bullet_tool(sku_list, uid_list, f_id_list, file_name)
                sg.popup("Run Complete!")
            except:
                sg.popup("Something went wrong. Please make sure everything was entered correctly.")

    bullet_window.close()

def make_converter_window():
    
    # Theme of windows
    sg.theme('Dark Grey 13')

    converter_layout = [[sg.Text("URL Converter Tool")],
                        [sg.Text("Be sure to close the Excel file being used BEFORE running the tool.", text_color='red', font=('Arial Bold', 10))],
                        [sg.Text("Enter Skus and URLs for Primary images:")],
                        [sg.Text('Please enter Primary Sku(MFG Number) list.'), sg.InputText(key='-SKU-', pad=(0,0))],
                        [sg.Text('Please enter Primary image URL list.'), sg.InputText(key='-URL-', pad=(0,0))],
                        [sg.Text("Enter Skus and URLs for Second image:")],
                        [sg.Text('Please enter Sku(MFG Number) list for second image.'), sg.InputText(key='-SKU2-', pad=(0,0))],
                        [sg.Text('Please enter image URL list for second image.'), sg.InputText(key='-URL2-', pad=(0,0))],
                        [sg.Text("Enter Skus and URLs for Third image:")],
                        [sg.Text('Please enter Sku(MFG Number) list for third image.'), sg.InputText(key='-SKU3-', pad=(0,0))],
                        [sg.Text('Please enter image URL list for third image.'), sg.InputText(key='-URL3-', pad=(0,0))],
                        [sg.Text("Enter Skus and URLs for Fourth image:")],
                        [sg.Text('Please enter Sku(MFG Number) list for fourth image.'), sg.InputText(key='-SKU4-', pad=(0,0))],
                        [sg.Text('Please enter image URL list for fourth image.'), sg.InputText(key='-URL4-', pad=(0,0))],
                        [sg.Text('Please enter the absolute path of Excel file to use.'), sg.InputText(key='-E_NAME-')],
                        [sg.Text('Please enter the absolute path of folder to download images to.'), sg.InputText(key='-F_NAME-')],
                        [sg.Button("Run"), sg.Button("Exit")]]
    
    convert_window = sg.Window('Bullet Scraper Window', converter_layout, modal=True)
    
    while True:
        
        event, values = convert_window.read()
        
        if event in(sg.WIN_CLOSED, "Exit"):
            break
        
        mfg_list_primary = values['-SKU-'].split('\n')
        mfg_list_2 = values['-SKU2-'].split('\n')
        mfg_list_3 = values['-SKU3-'].split('\n')
        mfg_list_4 = values['-SKU4-'].split('\n')
        Primary_list = values['-URL-'].split('\n')
        img_2_list = values['-URL2-'].split('\n')
        img_3_list = values['-URL3-'].split('\n')
        img_4_list = values['-URL4-'].split('\n')
        excel_file_name = r'{}'.format(values['-E_NAME-'].rstrip())
        folder_name = r'{}'.format(values['-F_NAME-'].rstrip())

        if event == 'Run':
            
            try:
                converter_tool(mfg_list_primary, mfg_list_2, mfg_list_3, mfg_list_4, Primary_list, img_2_list, img_3_list, img_4_list, excel_file_name, folder_name)
                # sg.popup("Run Complete!")
            except:
                sg.popup("Something went wrong. Please make sure everything was entered correctly.")

    convert_window.close()
    
def main():
    # Theme of windows
    sg.theme('Dark Grey 13')
    
    # Creating window layouts
    main_layout = [[sg.Text("Asset Scraper Tool")], 
                    [sg.Text("Be sure to close the Excel file being used BEFORE running the tool.", text_color='red', font=('Arial Bold', 10))],
                    [sg.Text("Choose which tool you want.")],
                    [sg.Button("Image URL Converter"), sg.Button("Image Scraper Tool"), sg.Button("Bullet Scraper Tool"), sg.Button("Exit")]]

    main_window = sg.Window('Main Window', main_layout)

    # Event Loop
    while True:
        event, values = main_window.read()

        
        # End program if conditions met
        if event in(sg.WIN_CLOSED, "Exit"):
                break
        
        # Runs the Image scraper tool window and tool
        elif event == 'Image Scraper Tool':
            make_img_window()
        
        # Runs the Bullet scraper tool window and tool
        elif event == 'Bullet Scraper Tool':
            make_bullet_window()

        # Runs the Image URL Converter tool window and tool
        elif event == 'Image URL Converter':
            make_converter_window()
    
    main_window.close()

# Run the program
if __name__ == "__main__":
    main()
    
